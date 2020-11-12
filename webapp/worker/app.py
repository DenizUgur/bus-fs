from openpyxl import load_workbook
import sys
import os

if __name__ == "__main__":
    # Get request info
    temp_file = sys.argv[1]
    student_id = sys.argv[2]

    # Template specific constant
    key_sheet = "KEY"
    id_cell = "A1"

    # Get the template file
    wb = load_workbook("../data/templates/{}_temp.xlsm".format(temp_file), keep_vba=True)

    # Get the KEY sheet
    ws = wb[key_sheet]

    # Protect the KEY sheet
    ws.protection.sheet = True
    ws.protection.password = '#!...'
    ws.protection.enable()
    ws.sheet_state = 'veryHidden'

    # Add the student_id to the sheet
    ws[id_cell] = student_id

    # Create /out folder if not exists
    if not os.path.isdir("../data/out"):
        os.mkdir("../data/out", 0o777)

    # Save the template file to /out
    wb.save("../data/out/{}_{}.xlsm".format(student_id, temp_file))